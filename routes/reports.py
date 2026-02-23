from flask import Blueprint, send_file, request, render_template, flash, redirect, url_for
from flask_login import login_required, current_user
from models.ticket import Ticket
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os

reports_bp = Blueprint("reports", __name__)


@reports_bp.route("/report", methods=["GET", "POST"])
@login_required
def report():
    if current_user.role not in ("ADMIN", "SUPER_ADMIN"):
        return "Unauthorized", 403

    # Default: previous calendar month
    today = datetime.utcnow()
    first_of_this_month = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    default_end = (first_of_this_month - timedelta(days=1)).replace(hour=23, minute=59, second=59)
    default_start = default_end.replace(day=1, hour=0, minute=0, second=0)

    if request.method == "POST":
        try:
            date_from = datetime.strptime(request.form["date_from"], "%Y-%m-%d")
            date_to = datetime.strptime(request.form["date_to"], "%Y-%m-%d").replace(
                hour=23, minute=59, second=59
            )
        except (ValueError, KeyError):
            flash("Invalid date range provided.", "danger")
            return redirect(url_for("reports.report"))

        if date_from > date_to:
            flash("Start date must be before end date.", "warning")
            return redirect(url_for("reports.report"))

        # Build query
        query = Ticket.query.filter(
            Ticket.created_at >= date_from,
            Ticket.created_at <= date_to
        )

        if current_user.role == "ADMIN":
            query = query.filter_by(department=current_user.department)

        tickets = query.all()

        # Build Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Ticket Report"

        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2563EB")

        headers = [
            "ID", "Title", "Issue Type", "Priority", "Status",
            "Department", "Location", "Raised By",
            "Created At", "SLA Due", "Closed At", "Acknowledged At"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for ticket in tickets:
            ws.append([
                ticket.id,
                ticket.title,
                ticket.issue_type,
                ticket.priority,
                ticket.status,
                ticket.department,
                ticket.location,
                ticket.created_by.username if ticket.created_by else "—",
                ticket.created_at.strftime("%Y-%m-%d %H:%M") if ticket.created_at else "",
                ticket.sla_due_at.strftime("%Y-%m-%d %H:%M") if ticket.sla_due_at else "",
                ticket.closed_at.strftime("%Y-%m-%d %H:%M") if ticket.closed_at else "",
                ticket.acknowledged_at.strftime("%Y-%m-%d %H:%M") if ticket.acknowledged_at else "",
            ])

        # Auto column width
        for col in ws.columns:
            max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        os.makedirs("instance", exist_ok=True)
        dept_label = current_user.department if current_user.role == "ADMIN" else "ALL"
        filename = f"ticket_report_{dept_label}_{date_from.strftime('%Y%m%d')}_to_{date_to.strftime('%Y%m%d')}.xlsx"
        filepath = os.path.join("instance", filename)
        wb.save(filepath)

        return send_file(filepath, as_attachment=True)

    # GET — render the form
    return render_template(
        "report.html",
        default_start=default_start.strftime("%Y-%m-%d"),
        default_end=default_end.strftime("%Y-%m-%d"),
    )
